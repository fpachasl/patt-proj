# apps/projects/views.py
from openpyxl import Workbook
from rest_framework import viewsets
from datetime import date, datetime
from django.db import models
from django.http import HttpResponse
from apps.company.models import Company
from apps.company.serializer import CompanySerializer
from apps.documents.models import Document
from apps.users.models import User
from .models import AreaLeader, Project, ProjectArea, ProjectAreaMember, ProjectLeader, ProjectPriority, ProjectState, ProjectType
from rest_framework.decorators import action
from .serializers import AreaLeaderSerializer, ProjectAreaMemberSerializer, ProjectAreaSerializer, ProjectDetailSerializer, ProjectLeaderSerializer, ProjectPrioritySerializer, ProjectSerializer,ProjectStateSerializer, ProjectTypeSerializer, UserSimpleSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils.timezone import now
from django.db.models import Count
from django.db.models.functions import TruncMonth
from collections import defaultdict
from apps.tasks.models import Task
from rest_framework import status
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectDetailSerializer
    permission_classes = [IsAuthenticated]
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProjectSerializer  # usa IDs planos para escribir
        return ProjectDetailSerializer  # usa serializers anidados para leer
    
    @action(detail=False, methods=['get'], url_path='dashboard')
    def get_dashboard(self, request):
        user = request.user

        if user.role.id == 1:
            projects = Project.objects.all()
        elif user.role.id == 2:
            projects = Project.objects.filter(leaders__leader=user)
        elif user.role.id == 3:
            # Empleado: proyectos donde está asignado como miembro de área
            projects = Project.objects.filter(
                areas__members__user=user
            ).distinct()
            
            # Tareas asignadas al miembro
            tasks = Task.objects.filter(assigned_user=user)

            return Response({
                "projects_assigned": projects.count(),
                "pending_tasks": tasks.filter(task_state__code="planning").count(),
                "tasks_in_progress": tasks.filter(task_state__code="in_progress").count(),
                "tasks_completed": tasks.filter(task_state__code="completed").count(),
            })
        
        else:
            return Response({"detail": "No autorizado para ver esta información."}, status=403)

        data = {
            "total": projects.count(),
            "planning": projects.filter(project_state__code="planning").count(),
            "in_progress": projects.filter(project_state__code="in_progress").count(),
            "completed": projects.filter(project_state__code="completed").count(),
            "on_hold": projects.filter(project_state__code="on_hold").count(),
            "cancelled": projects.filter(project_state__code="cancelled").count(),
            "delayed": projects.filter(end_date__lt=now().date(), project_state__code__in=["planning", "in_progress"]).count()
        }

        return Response(data)
    
    @action(detail=False, methods=['get'], url_path='dashboard-chart')
    def get_dashboard_chart(self, request):
        user = request.user

        # Filtra proyectos visibles según rol
        if user.role.id == 1:
            visible_projects = Project.objects.all()
        elif user.role.id == 2:
            visible_projects = Project.objects.filter(leaders__leader=user)
        else:
            return Response({"detail": "No autorizado"}, status=403)

        # Agrupa proyectos por mes
        project_by_month = (
            visible_projects
            .annotate(month=TruncMonth("start_date"))
            .values("month")
            .annotate(projects=Count("id"))
            .order_by("month")
        )

        # Agrupa tareas por mes de inicio
        from apps.tasks.models import Task
        task_by_month = (
            Task.objects.filter(project_id__in=visible_projects.values_list("id", flat=True))
            .annotate(month=TruncMonth("start_date"))
            .values("month")
            .annotate(tasks=Count("id"))
            .order_by("month")
        )

        # Mezcla ambos resultados
        result = defaultdict(lambda: {"projects": 0, "tasks": 0})

        for item in project_by_month:
            month = item["month"].strftime("%B")
            result[month]["projects"] = item["projects"]

        for item in task_by_month:
            month = item["month"].strftime("%B")
            result[month]["tasks"] = item["tasks"]

        final_data = [
            {"month": month, "projects": result[month]["projects"], "tasks": result[month]["tasks"]}
            for month in result
        ]

        return Response(final_data)
    
    @action(detail=False, methods=["get"], url_path="assigned")
    def assigned_projects(self, request):
        user = request.user

        if not user.is_authenticated:
            return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)

        # Si es admin, ve todos
        if user.role.id == 1:
            projects = Project.objects.all()

        # Si es líder, ve los que lidera y los donde es miembro
        elif user.role.id == 2:
            projects = Project.objects.filter(
                models.Q(leaders__leader=user) |
                models.Q(areas__members__user=user)
            ).distinct()

        # Si es miembro, solo los donde es miembro
        else:
            projects = Project.objects.filter(
                areas__members__user=user
            ).distinct()

        serializer = ProjectDetailSerializer(projects, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=["put", "patch"], url_path="update-project")
    def update_project(self, request):
        user = request.user

        if not user.is_authenticated:
            return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)

        if user.role.id not in [1, 2]:  # Solo admin o líder
            return Response({
                "detail": "No tiene permisos para actualizar el proyecto",
                "status": "forbidden"
            }, status=status.HTTP_403_FORBIDDEN)

        project_id = request.data.get("project_id")
        if not project_id:
            return Response({"detail": "ID del proyecto requerido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"detail": "Proyecto no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        # Actualizar campos si existen en request
        if "name" in request.data:
            project.name = request.data["name"]

        if "description" in request.data:
            project.description = request.data["description"]

        if "end_date" in request.data:
            project.end_date = request.data["end_date"]

        if "project_type" in request.data:
            try:
                pt = ProjectType.objects.get(id=request.data["project_type"])
                project.project_type = pt
            except ProjectType.DoesNotExist:
                return Response({"detail": "Tipo de proyecto inválido"}, status=status.HTTP_400_BAD_REQUEST)

        if "project_state" in request.data:
            try:
                ps = ProjectState.objects.get(id=request.data["project_state"])
                project.project_state = ps
            except ProjectState.DoesNotExist:
                return Response({"detail": "Estado del proyecto inválido"}, status=status.HTTP_400_BAD_REQUEST)

        if "company" in request.data:
            try:
                company = Company.objects.get(id=request.data["company"])
                project.company = company
            except Company.DoesNotExist:
                return Response({"detail": "Empresa inválida"}, status=status.HTTP_400_BAD_REQUEST)

        if "priority" in request.data:
            try:
                priority = ProjectPriority.objects.get(id=request.data["priority"])
                project.priority = priority
            except ProjectPriority.DoesNotExist:
                return Response({"detail": "Prioridad inválida"}, status=status.HTTP_400_BAD_REQUEST)

        project.save()

        serializer = ProjectDetailSerializer(project)
        return Response({
            "detail": "Proyecto actualizado correctamente",
            "data": serializer.data,
            "status": "success"
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='export-report')
    def export_project_report(self, request):
        user = request.user

        if not user.is_authenticated:
            return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)

        # Filtrar proyectos según el rol del usuario
        if user.role.id == 1:  # Admin
            projects = Project.objects.all()
        elif user.role.id == 2:  # Líder
            projects = Project.objects.filter(leaders__leader=user).distinct()
        else:
            return Response({"message": "Solo el líder o el admin puede generar reportes"}, status=status.HTTP_403_FORBIDDEN)

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Proyectos"

        # Encabezados
        headers = [
            "Nombre", "Estado", "Tipo", "Prioridad", "Empresa",
            "Fecha inicio", "Fecha fin", "Líder(es)", "Tareas asignadas"
        ]
        ws.append(headers)

        # Aplicar estilo a encabezados
        header_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Llenar datos
        for p in projects:
            leaders_names = ", ".join([f"{l.leader.first_name} {l.leader.last_name}" for l in p.leaders.all()])
            task_count = Task.objects.filter(project=p).count()

            row = [
                p.name,
                p.project_state.name if p.project_state else "",
                p.project_type.name if p.project_type else "",
                p.priority.name if p.priority else "",
                p.company.name if p.company else "",
                p.start_date.strftime("%d/%m/%Y") if p.start_date else "",
                p.end_date.strftime("%d/%m/%Y") if p.end_date else "",
                leaders_names,
                task_count,
            ]
            ws.append(row)

            # Si el proyecto está retrasado, colorear de rojo
            if p.end_date and p.end_date < date.today() and p.project_state.code != "completed":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        # Preparar respuesta
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=reporte_proyectos.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-kpis')
    def export_project_kpis(self, request):
        user = request.user
        if not user.is_authenticated:
            return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)

        # 1. Filtrar proyectos según el rol
        if user.role.id == 1:  # Admin
            projects = Project.objects.all()
        elif user.role.id == 2:  # Líder
            projects = Project.objects.filter(leaders__leader=user).distinct()
        else:
            return Response({"detail": "Solo líderes o admins pueden generar reportes"}, status=status.HTTP_403_FORBIDDEN)

        total_projects = projects.count()
        project_ids = projects.values_list('id', flat=True)

        total_tasks = Task.objects.filter(project_id__in=project_ids).count()
        total_documents = Document.objects.filter(project_id__in=project_ids).count()
        total_users = User.objects.filter(is_active=True).count()
        completed_projects = projects.filter(project_state__code="completed").count()
        late_projects = projects.filter(end_date__lt=date.today()).exclude(project_state__code="completed").count()
        no_task_projects = projects.annotate(task_count=models.Count('tasks')).filter(task_count=0).count()
        critical_projects = projects.filter(end_date__lt=date.today(), project_state__code__in=["planning", "in_progress"])
        next_due_project = projects.filter(end_date__gte=date.today()).order_by("end_date").first()

        # 2. Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen KPIs"

        def style_header(row):
            fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = fill

        # 3. Información general
        ws.append(["Reporte de KPIs - Gestión de Proyectos"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.append(["Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")])
        ws.append(["Generado por", f"{user.first_name} {user.last_name} ({user.role.name})"])
        ws.append([])

        # 4. KPIs generales
        ws.append(["Indicador", "Valor"])
        style_header(ws[ws.max_row])
        ws.append(["Total proyectos", total_projects])
        ws.append(["Total tareas", total_tasks])
        ws.append(["Total documentos", total_documents])
        ws.append(["Usuarios activos", total_users])
        ws.append(["Proyectos completados (%)", f"{(completed_projects / total_projects) * 100:.2f}%" if total_projects else "0%"])
        ws.append(["Proyectos retrasados", late_projects])
        ws.append(["Proyectos sin tareas asignadas", no_task_projects])
        ws.append(["Promedio tareas por proyecto", round(total_tasks / total_projects, 2) if total_projects else 0])
        ws.append(["Promedio documentos por proyecto", round(total_documents / total_projects, 2) if total_projects else 0])
        ws.append([])

        # 5. Proyecto más próximo a vencer
        ws.append(["Proyecto más próximo a vencer"])
        style_header(ws[ws.max_row])
        if next_due_project:
            ws.append(["Nombre", next_due_project.name])
            ws.append(["Fecha de entrega", next_due_project.end_date.strftime("%d/%m/%Y")])
            ws.append(["Estado actual", next_due_project.project_state.name])
        else:
            ws.append(["Sin proyectos con fechas futuras."])
        ws.append([])

        # 6. Proyectos críticos
        ws.append(["Proyectos críticos (retrasados y no completados)"])
        style_header(ws[ws.max_row])
        if critical_projects.exists():
            ws.append(["Nombre", "Fecha fin", "Estado"])
            for cp in critical_projects:
                ws.append([
                    cp.name,
                    cp.end_date.strftime("%d/%m/%Y") if cp.end_date else "Sin fecha",
                    cp.project_state.name if cp.project_state else "Desconocido"
                ])
        else:
            ws.append(["Ningún proyecto crítico en este momento."])

        # 7. Estilo ancho de columnas
        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 35

        # 8. Respuesta
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename=KPIs_proyectos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='members')
    def get_project_members(self, request, pk=None):
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({"detail": "Proyecto no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        # Buscar todos los miembros del proyecto (desde las áreas)
        users = User.objects.filter(area_memberships__area__project=project).distinct()
        
        print(users)
        serializer = UserSimpleSerializer(users, many=True)
        return Response(serializer.data)
    
class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [IsAuthenticated]    
    
class ProjectPriorityViewSet(viewsets.ModelViewSet):
    queryset = ProjectPriority.objects.all()
    serializer_class = ProjectPrioritySerializer
    permission_classes = [IsAuthenticated]


class ProjectTypeViewSet(viewsets.ModelViewSet):
    queryset = ProjectType.objects.all()
    serializer_class = ProjectTypeSerializer
    permission_classes = [IsAuthenticated]


class ProjectStateViewSet(viewsets.ModelViewSet):
    queryset = ProjectState.objects.all()
    serializer_class = ProjectStateSerializer
    permission_classes = [IsAuthenticated]


class ProjectLeaderViewSet(viewsets.ModelViewSet):
    queryset = ProjectLeader.objects.all()
    serializer_class = ProjectLeaderSerializer
    permission_classes = [IsAuthenticated]


class ProjectAreaViewSet(viewsets.ModelViewSet):
    queryset = ProjectArea.objects.all()
    serializer_class = ProjectAreaSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get("project")
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset

class ProjectAreaMemberViewSet(viewsets.ModelViewSet):
    queryset = ProjectAreaMember.objects.all()
    serializer_class = ProjectAreaMemberSerializer
    permission_classes = [IsAuthenticated]


class AreaLeaderViewSet(viewsets.ModelViewSet):
    queryset = AreaLeader.objects.all()
    serializer_class = AreaLeaderSerializer
    permission_classes = [IsAuthenticated]